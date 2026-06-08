r"""工艺卡片 Excel 生成器。

根据 manufacturing_process.json 的工序数据和 lens_calc 的计算结果，
生成参照 M0454-G10 格式的完整工艺卡片 Excel。

静态单元格布局由 export_layout.json 驱动（使用 ${xxx} 占位符），
运行时由 _make_ctx() 统一替换，与工序模板使用同一套取值逻辑。

字体: 等线  填充色: 黄#FFFF00 橙#FFC000
工序表头三色: A-E#D9E1F4  F-J#E3F2D9  K-M#DCE6F2
"""

import json, os, datetime, re
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

from set import Tasks, load_preset
from lens_calc import LensParams, CalcResult, calculate


# ═══ 加载布局 JSON ═══
def _load_layout() -> dict:
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "export_layout.json")
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

_LAYOUT = _load_layout()

# ═══ 加载 Schema（用于 _make_ctx） ═══
def _load_schema() -> dict:
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "field_schema.json")
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

_SCHEMA = _load_schema()

# ═══ 样式（openpyxl 对象，无法 JSON 化） ═══
_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 由 _LAYOUT["cells"] 的 style 字段引用
_STYLE_MAP = {
    "title":  Font(name="等线", size=16, bold=True),
    "sec":    Font(name="等线", size=14, bold=True),
    "lbl":    Font(name="等线", size=12, bold=True),
    "bold11": Font(name="等线", size=11, bold=True),
    "bold10": Font(name="等线", size=10, bold=True),
    "11":     Font(name="等线", size=11),
}

# 由 _LAYOUT["cells"] 的 fill 字段引用
_FILL_MAP = {
    "h1":     PatternFill(start_color="D9E1F4", end_color="D9E1F4", fill_type="solid"),
    "h2":     PatternFill(start_color="E3F2D9", end_color="E3F2D9", fill_type="solid"),
    "h3":     PatternFill(start_color="DCE6F2", end_color="DCE6F2", fill_type="solid"),
    "yellow": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "orange": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
}
_FILL_NONE = PatternFill(fill_type=None)

_BAY_C = {"成形": "00B0F0", "球面": "FFC000", "平面": "FFC000",
          "清洗": "FF0000", "镀膜": "00B050"}


def _irr_rms_str(irr: float, rms: float) -> str:
    """面形列：IRR 和 RMS 条件组合。值为 0 时跳过。"""
    parts = []
    if irr:
        parts.append(f"IRR:{irr}\u03bb")
    if rms:
        parts.append(f"RMS:{rms:.0f}nm")
    return ", ".join(parts)


def _cell(ws, ref, val, font=None, fill=None, align=_ALIGN_C, border=_BORDER_ALL):
    """写入单元格。ref 为 Excel 引用如 'B3' 或 'AB12'。"""
    col = openpyxl.utils.cell.column_index_from_string(ref.rstrip("0123456789"))
    row = int("".join(c for c in ref if c.isdigit()))
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or _STYLE_MAP["bold11"]
    c.fill = fill or _FILL_NONE
    c.alignment = align
    c.border = border
    return c


def _bay_fill(bay):
    for p, c in _BAY_C.items():
        if bay.startswith(p):
            return PatternFill(start_color=c, end_color=c, fill_type="solid")
    return None


def _make_ctx(p, r):
    """从 field_schema.json export_ctx 生成占位符 → 值的映射字典。

    这是整个系统的唯一取值入口 —— 工序模板和 Excel 布局共用同一份 ctx。
    """
    ctx = {}
    for item in _SCHEMA["export_ctx"]:
        if "ctx" not in item:
            continue
        src = p if item["source"] == "params" else r
        val = getattr(src, item["attr"])
        if item.get("hide_zero"):
            try:
                if float(val) == 0.0:
                    ctx[item["ctx"]] = ""
                    continue
            except (ValueError, TypeError):
                if not val:
                    ctx[item["ctx"]] = ""
                    continue
        text = str(val) if item["fmt"] == "s" else _remove_trailing_zeros(format(val, item["fmt"]))
        if item.get("prefix"):
            text = item["prefix"] + text
        if item.get("suffix"):
            text += item["suffix"]
        ctx[item["ctx"]] = text
    return ctx


def _remove_trailing_zeros(text: str) -> str:
    """移除数值字符串末尾的零和小数点。如 '48.0' → '48', '14.890' → '14.89'。"""
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def _resolve(val: str, ctx: dict) -> str:
    """将字符串中的 ${xxx} 占位符替换为实际值。"""
    for k, v in ctx.items():
        val = val.replace(k, v)
    return val


# 匹配 {last_row+N} 相对行号语法
_RE_LAST_ROW = re.compile(r'\{last_row\+(\d+)\}')

def _resolve_row(text: str, last_row: int) -> str:
    """将字符串中的 {last_row+N} 替换为实际行号。"""
    def _repl(m):
        return str(last_row + int(m.group(1)))
    return _RE_LAST_ROW.sub(_repl, text)


def _render_footer(ws, ctx, last_row):
    """由 _LAYOUT["footer"] 驱动渲染页脚。"""
    footer = _LAYOUT.get("footer")
    if not footer:
        return
    ctx = dict(ctx)
    today = datetime.date.today()
    ctx["${today}"] = f"{today.year}.{today.month}.{today.day}"
    for row_def in footer.get("rows", []):
        if not row_def["cells"]:
            continue
        ref0 = _resolve_row(row_def["cells"][0]["ref"], last_row)
        row_num = int("".join(c for c in ref0 if c.isdigit()))
        if "row_height" in row_def:
            ws.row_dimensions[row_num].height = row_def["row_height"]
        for cd in row_def["cells"]:
            ref = _resolve_row(cd["ref"], last_row)
            raw = cd["value"]
            val = _resolve(raw, ctx)
            style_key = cd.get("style", "bold11")
            font = _STYLE_MAP.get(style_key, _STYLE_MAP["bold11"])
            merge = cd.get("merge")
            if merge:
                ws.merge_cells(_resolve_row(merge, last_row))
            _cell(ws, ref, val, font=font)
        for m in row_def.get("merges", []):
            ws.merge_cells(_resolve_row(m, last_row))


def _apply_all_borders(ws):
    """对工作表中所有单元格重新应用完整边框，解决合并区域边框丢失的问题。"""
    if ws.max_row is None or ws.max_column is None:
        return
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_ALL


# ═══ 主入口 ═══
def generate_process_card(filepath: str, *, lens_params=None, calc_result=None):
    p = lens_params or LensParams()
    r = calc_result or calculate(p)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _LAYOUT.get("sheet_name", "工艺卡")
    _cl = get_column_letter

    # ── 列宽（来自布局 JSON） ──
    for col_letter, width in _LAYOUT.get("column_widths", {}).items():
        ci = openpyxl.utils.cell.column_index_from_string(col_letter)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── 行高（来自布局 JSON） ──
    for row_str, height in _LAYOUT.get("row_heights", {}).items():
        ws.row_dimensions[int(row_str)].height = height

    # ── 行1-12: 静态单元格（来自布局 JSON + 占位符替换） ──
    ctx = _make_ctx(p, r)
    for celldef in _LAYOUT.get("cells", []):
        ref = celldef["ref"]
        raw = celldef["value"]
        val = _resolve(raw, ctx)
        style_key = celldef.get("style", "bold11")
        font = _STYLE_MAP.get(style_key, _STYLE_MAP["bold11"])
        fill_key = celldef.get("fill")
        fill = _FILL_MAP.get(fill_key) if fill_key else None
        merge = celldef.get("merge")
        if merge:
            ws.merge_cells(merge)
        _cell(ws, ref, val, font=font, fill=fill)

    # ── 额外合并范围（布局 JSON 中的 merge_ranges） ──
    for mr in _LAYOUT.get("merge_ranges", []):
        ws.merge_cells(mr)

    # ── 动态单元格（需条件逻辑，不适合 JSON） ──
    # F4 / F5: R值 — R=0 时显示"平面"
    _cell(ws, "F4", _remove_trailing_zeros(f"{p.r1:.3f}") if p.r1 else "平面", font=_STYLE_MAP["bold11"])
    _cell(ws, "F5", _remove_trailing_zeros(f"{p.r2:.3f}") if p.r2 else "平面", font=_STYLE_MAP["bold11"])
    # J4 / J5: 面形 — 条件组合 IRR + RMS
    _cell(ws, "J4", _irr_rms_str(p.s1_irr, p.s1_rms), font=_STYLE_MAP["bold11"])
    _cell(ws, "J5", _irr_rms_str(p.s2_irr, p.s2_rms), font=_STYLE_MAP["bold11"])

    # ── 行13+: 工序数据（动态循环，保留原逻辑） ──
    seq, row, bf = 0, 13, {}
    for g in Tasks:
        bay = g.bay.strip()
        proc = g.process.strip()
        obj = g.obj.strip() if g.obj else ""
        fill = bf.setdefault(bay, _bay_fill(bay))
        seq += 1
        gs = row
        for i, rt in enumerate(g.requires):
            req = rt.strip()
            req = _resolve(req, ctx)  # 统一走占位符替换
            _cell(ws, f"A{row}", seq, font=_STYLE_MAP["bold11"], fill=fill)
            _cell(ws, f"B{row}", bay if i == 0 else "", font=_STYLE_MAP["bold11"], fill=fill)
            _cell(ws, f"C{row}", proc if i == 0 else "", font=_STYLE_MAP["bold11"], fill=fill)
            _cell(ws, f"D{row}", obj if i == 0 else "", font=_STYLE_MAP["bold11"], fill=fill)
            _cell(ws, f"E{row}", req, font=_STYLE_MAP["11"], align=_ALIGN_C)
            for c in range(6, 14):
                _cell(ws, f"{_cl(c)}{row}", "", font=_STYLE_MAP["bold11"])
            ws.row_dimensions[row].height = getattr(g, "row_height", None) or max(20, 14 * max(1, len(req) // 40))
            row += 1
        ge = row - 1
        if ge > gs:
            for c in (1, 2, 3, 4):
                ws.merge_cells(start_row=gs, start_column=c, end_row=ge, end_column=c)
    last_row = row - 1

    # ── 工序区纵向合并 ──
    ws.merge_cells(start_row=13, start_column=13, end_row=last_row, end_column=15)
    for mr in range(13, last_row + 1):
        ws.merge_cells(start_row=mr, start_column=5, end_row=mr, end_column=7)
    for mr in range(12, last_row + 1):
        ws.merge_cells(start_row=mr, start_column=8, end_row=mr, end_column=9)

    # ── O4: 单位 "mm" + 合并（先写后 merge，避免 MergedCell 报错） ──
    _cell(ws, "O4", "mm", font=_STYLE_MAP["bold11"])
    ws.merge_cells("O4:O5")

    # ── 页脚（由 _LAYOUT["footer"] + {last_row+N} 驱动） ──
    _render_footer(ws, ctx, last_row)

    # ── 统一重绘边框（解决合并区域边框丢失） ──
    _apply_all_borders(ws)

    # ── 页面设置 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = None

    wb.save(filepath)
    return ws.max_row


def export_process_card(filepath: str, lens_params=None, preset_path="manufacturing_process.json"):
    if os.path.exists(preset_path):
        load_preset(preset_path)
    p = lens_params or LensParams()
    r = calculate(p)
    generate_process_card(filepath, lens_params=p, calc_result=r)
    return r
